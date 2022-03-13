from openpyxl import load_workbook

wb = load_workbook("tp_pistons.xlsx")
sheet = wb[wb.sheetnames[0]]


def cleanValue(value):
    return float(value)


def cleanValueInt(value):
    return int(value)


t = [cleanValue(cell.value) for cell in sheet["A"][1:]]
v_veh = [cleanValue(cell.value) for cell in sheet["B"][1:]]
pente = [cleanValue(cell.value) for cell in sheet["C"][1:]]
rapport = [cleanValueInt(cell.value) for cell in sheet["D"][1:]]
q_carb_mgcp = [cleanValue(cell.value) for cell in sheet["E"][1:]]

# t [s] : temps
# v_veh [km/h] : vitesse du véhicule
# pente [°] : pente de la route, en degrés
# rapport [-] : rapport de boite engagé
# q_carb [mg/cp] : débit de carburant (mg/cp : milligrammes par coup, masse de carburant injectée par cylindre pour 1 cycle thermodynamique)


# delta_t [s] Ecart entre deux prises de mesures
delta_t = t[1] - t[0]

assert len(t) == len(v_veh) == len(pente) == len(rapport) == len(q_carb_mgcp)
nbEtapes = len(t)
print(f"Chargé {nbEtapes} lignes du fichier")
